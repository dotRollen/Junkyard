#! python2

import subprocess
import os
import platform
import re
import time
import sys
import argparse
from openpyxl import Workbook

def ping(host):

    ping_str = "-n 1" if platform.system().lower() == "windows" else "-c 1"
    args = "ping " + " " + ping_str + " " + host
    return subprocess.call(args, shell=need_sh(), stdout=FNULL(), stderr=subprocess.STDOUT) == 0


def need_sh():

    return False if platform.system().lower() == "windows" else True


def FNULL():

    return open(os.devnull, 'w')

def getinv(hosts):
	wb = Workbook()
	macHeaders = ['Device Name', 'Model', 'Service Tag', 'iDRAC NIC MAC', '1G NIC.Integrated.1-1-1 MAC', 'NIC.Integrated.1-2-1', 
				'NIC.Integrated.1-3-1 MAC', 'NIC.Integrated.1-4-1 MAC', 'NIC.Slot*-1-1 MAC', 'NIC.Slot.*-2-1 MAC']
	fwHeaders = ['Device Name', 'BIOS Firmware', 'RAID Firmware', 'NIC 1-1-1 Firwmare', 'Enclosure Firmware', 'iDRAC Embedded Firmware', 'NIC.Slot*-1-1 Firmware', 'NIC.Slot.*-2-1 Firmware']
	ws1 = wb.active
	ws1.title = "MAC Address"
	ws2 = wb.create_sheet("Firmware")
	ws3 = wb.create_sheet("BIOS Settings")
	ws1.append(macHeaders) 
	ws2.append(fwHeaders)
	#ws3.append(biosHeaders)
	for host in hosts:
		folder = (re.findall('\d+', host ))
		if ping(host):
			try:
				print("Collecting hwinventory for..." + host)
				subprocess.call("mkdir -p /output/GAIA/rack" + folder[0] + "/u" + folder[1] + "/; racadm -r " + host + " -u root -p calvin hwinventory > /output/GAIA/rack" + folder[0] + "/u" + folder[1] + "/" + host + "_hwinventory.txt", shell=need_sh(), stdout=FNULL(), stderr=subprocess.STDOUT)
				#print("Collecting BIOS configuration for..." + host)
				#subprocess.call("mkdir -p /output/GAIA/rack" + folder[0] + "/u" + folder[1] + "/;racadm -r " + host + " -u root -p calvin get -t xml -f /output/GAIA/rack" + folder[0] + "/u" + folder[1] + "/" + host + "_biosconfig.xml", shell=need_sh())
				hwinv = open("/output/GAIA/rack" + folder[0] + "/u" + folder[1] + "/" + host + "_hwinventory.txt")
				#xmlconfig = open("/output/GAIA/rack" + folder[0] + "/u" + folder[1] + "/" + host + "_biosconfig.xml")
				searchlines = hwinv.readlines()
				#searchlines += xmlconfig.readlines()
				hwinv.close()
				print("Collecting inventory data for..." + host)
				for i, line in enumerate(searchlines):
					if "BIOSVersionString = " in line: 
						fwBios = line.replace('BIOSVersionString = ', '')
					elif "ControllerFirmwareVersion = " in line: 
						fwRaid = line.replace('ControllerFirmwareVersion = ', '')
					elif "[InstanceID: NIC.Integrated.1-1-1]" in line and "Device Type = NIC" in searchlines[i+1]:
						for l in searchlines[i:i+32]:
							if 'FamilyVersion = ' in l:
								fwNic1 = l.replace('FamilyVersion = ', '')
							if "PermanentMACAddress = " in l:
								nic1MAC = l.replace('PermanentMACAddress = ', '')
					elif "[InstanceID: NIC.Integrated.1-2-1]" in line and "Device Type = NIC" in searchlines[i+1]:
						for l in searchlines[i:i+32]:
							if 'FamilyVersion = ' in l:
								fwNic2 = l.replace('FamilyVersion = ', '')
							if "PermanentMACAddress = " in l:
								nic2MAC = l.replace('PermanentMACAddress = ', '')
					elif "[InstanceID: NIC.Integrated.1-3-1]" in line and "Device Type = NIC" in searchlines[i+1]:
						for l in searchlines[i:i+32]:
							if 'FamilyVersion = ' in l:
								fwNic3 = l.replace('FamilyVersion = ', '')
							if "PermanentMACAddress = " in l:
								nic3MAC = l.replace('PermanentMACAddress = ', '')
					elif "[InstanceID: NIC.Integrated.1-4-1]" in line and "Device Type = NIC" in searchlines[i+1]:
						for l in searchlines[i:i+32]:
							if 'FamilyVersion = ' in l:
								fwNic4 = l.replace('FamilyVersion = ', '')
							if "PermanentMACAddress = " in l:
								nic4MAC = l.replace('PermanentMACAddress = ', '')
					elif "[InstanceID: Enclosure.Internal.0-1:RAID.Integrated.1-1]" in line: 
						for l in searchlines[i:i+13]:
							if 'Version = ' in l:
								fwEnclosure = l.replace('Version = ', '')
					elif "[InstanceID: iDRAC.Embedded.1-1#IDRACinfo]" in line:
						for l in searchlines[i:i+13]:
							if 'FirmwareVersion = ' in l:
								fwIdrac = l.replace('FirmwareVersion = ', '')	
							if "PermanentMACAddress = " in l:
								idracMAC = l.replace('PermanentMACAddress = ', '')
					elif "[InstanceID: System.Embedded.1]" in line:
						for l in searchlines[i:i+56]:
							if "ServiceTag = " in l:
								serviceTAG = l.replace('ServiceTag = ', '')
							if "Model = " in l:
								modelOEM = l.replace('Model = ', '')
					elif ("[InstanceID: NIC.Slot.2-1-1]" in line or "[InstanceID: NIC.Slot.3-1-1]" in line) and "Device Type = NIC" in searchlines[i+1]:
						for l in searchlines[i:i+33]:
							if "FamilyVersion = " in l:
								nic1FW10G = l.replace('FamilyVersion = ', '')
							if "PermanentMACAddress = " in l:
								nic1MAC10G = l.replace('PermanentMACAddress = ', '')
					elif ("[InstanceID: NIC.Slot.2-2-1]" in line or "[InstanceID: NIC.Slot.3-2-1]" in line) and "Device Type = NIC" in searchlines[i+1]:								
						for l in searchlines[i:i+33]:
							if "FamilyVersion = " in l:
								nic2FW10G = l.replace('FamilyVersion = ', '')
							if "PermanentMACAddress = " in l:
								nic2MAC10G = l.replace('PermanentMACAddress = ', '')

			except (subprocess.CalledProcessError) as error:	
				print(error.output)
				sys.exit()
			finally:
				print("Appending data for " + host + " to worksheet")
				writeFW = [host, fwBios, fwRaid, fwNic1, fwEnclosure, fwIdrac, nic1FW10G, nic2FW10G]
				writeMAC = [host, modelOEM, serviceTAG, idracMAC, nic1MAC, nic2MAC, nic3MAC, nic4MAC, nic1MAC10G, nic2MAC10G ]
				#writeBIOS = [host, fwBios, fwRaid, fwNic, fwEnclosure, fwIdrac, tenG]
				ws1.append(writeMAC)
				ws2.append(writeFW)
				#ws3.append(writeBIOS)

		else:
			print("ERROR: Can not ping system " + host + "!!!")
	wb.save(filename=time.strftime("%Y%m%d-%H%M%S") + ".xlsx")
	print(host + " hardware inventory complete")

def bindbmc(hosts):
    for host in hosts:
        if ping(host):
            mac = subprocess.check_output("psh " + host + " ipmitool lan print 1|grep -io '[0-9A-F]\{2\}\(:[0-9A-F]\{2\}\)\{5\}' |egrep -m 1 '\w+'", shell=need_sh(), stderr=subprocess.STDOUT)
            subprocess.call("chdef " + host + "-bmc mac=" + mac, shell=need_sh(), stdout=FNULL(), stderr=subprocess.STDOUT)
            bmcmac = subprocess.check_output("lsdef " + host + "-bmc | grep 'mac=' | awk -F= '{print $NF}'", shell=need_sh(), stderr=subprocess.STDOUT)
            if mac == bmcmac:
                subprocess.call('makedhcp -a ' + host + "-bmc", shell=need_sh(), stdout=FNULL(), stderr=subprocess.STDOUT)	
                print("Complete: " + host + "-bmc MAC address has been bound.")
            else:
                print("ERROR: MAC address for " + host + "-bmc mac:" + mac)
        else:
            print("ERROR: Can not ping system " + host + "!!!")

    sys.exit()


def main():
    parser = argparse.ArgumentParser(description='Automate BMC mac binding, get hardware and software inventory')

    parser.add_argument("noderange", metavar="<noderange>", help='groups or node range IE: node1,node2,node3 | rack1')
    parser.add_argument("-g", "--getinv", action='store_true', help='get hardware and software inventory for Dell R630 and R730xd')
    parser.add_argument("-b", "--bmcbind", action='store_true', help='bind host\'s BMC mac to BMC host definition.')

    args = parser.parse_args()

    try:
        hosts = subprocess.check_output("nodels " + args.noderange, shell=True, stderr=subprocess.STDOUT).split()
		
        if args.getinv == True or args.bmcbind == True:

            if args.getinv:
                getinv(hosts)

            if args.bmcbind:
                bindbmc(hosts)

        else:
            print(
                "Error: Must include at least one argument. Type \"inv.py -h\" or --help for more information.")

    except (subprocess.CalledProcessError) as error:
        print(error.output)
        sys.exit()


if __name__ == "__main__":
    main()