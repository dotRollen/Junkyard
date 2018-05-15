#! python2

from openpyxl import load_workbook
import re

nodes = []

class Device(object):
    def __init__(self, deviceName, deviceModel, deviceSerial, deviceBmcMAC, deviceIntergratedNicMACs, device10gMACs):
        self.name = deviceName
        self.model = deviceModel
        self.serial = deviceSerial
        self.bmcMAC = deviceBmcMAC
        self.integratedNICMACs = deviceIntergratedNicMACs
        self.secondaryNICMACs = device10gMACs
        self.rack = re.search('(?<=cld)([0-9]{2})', self.name).group(0)
        self.u = re.search('(?<=u)(\d{2})', self.name).group(0)
        

    def __str__(self):
        return "Device Name: %s (Rack %s, Unit %s)\n\tModel: %s\n\tSerial: %s\n\tBMC MAC: %s\n\tIntegrated NIC MACs: %s\n\t10g NIC MACs: %s\n" % (self.name, self.rack, self.u, self.model, self.serial, 
                                                                                                                   self.bmcMAC, ', '.join(self.integratedNICMACs), ', '.join(self.secondaryNICMACs))

path = raw_input("Path to .xlsx file to be read: ")
wb = load_workbook(path)

row=2
while row < wb.active.max_row:
    if wb.active['A%d' % row].value != None:
       
        integratedNICs = [wb.active['E%d' % row].value, wb.active['F%d' % row].value, wb.active['G%d' % row].value, wb.active['H%d' % row].value]
        secondaryNICs = [wb.active['I%d' % row].value, wb.active['J%d' % row].value]
        nodes.append(Device(wb.active['A%d' % row].value, 
                            wb.active['B%d' % row].value, 
                            wb.active['C%d' % row].value, 
                            wb.active['D%d' % row].value, 
                            integratedNICs, 
                            secondaryNICs ))

    row+=1    
for device in nodes:
    print device
print 'Total Nodes: ' + str(row - 1)