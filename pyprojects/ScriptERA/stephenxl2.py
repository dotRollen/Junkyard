#!/usr/bin/python
#
#########################################################################
##                                                                     ##
##                          ScriptERA                                  ##
##                      Version 1.0 8/26/2016                          ##
##                                                                     ##
##        ----------------------***----------------------              ##
##                                                                     ##
## Change log:                                                         ##
##                                                                     ##
##                                                                     ##
##                                                                     ##
##                                                                     ##
##                                                                     ##
##        ----------------------***----------------------              ##
##                                                                     ##
##          .-~*`"-"`*~-.-(_Contributors_)-,.-~*`"-"`*~-.              ##
##                                                                     ##
##  Ed Darrah (Ed_Darrah@dell.com),                                    ##
##  Shawn Thepuatrakul (Shawn_Thepuatrakul@dell.com)                   ##
##  Ed Nunez (Ed_Nunez@shi.com)                                        ##
##  Stephen Mourad (StephenMourad711@gmail.com)                        ##
##                                                                     ##
##        .-~*`"-"`*~-.-(_Special Thanks To_)-,.-~*`"-"`*~-.           ##
##                                                                     ##
##                 Ed Darrah and Shawn Thepuatrakul                    ##
##       Originally created the program to run ping, error checking    ##
##                and automating auto-configuration                    ##
##                                                                     ##
#########################################################################
#
# Python Dependent Module Imports
import subprocess, os, platform, re, time

from openpyxl import Workbook
from openpyxl.compat import range
from openpyxl.cell import get_column_letter, Cell
from openpyxl.comments.comments import Comment
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font, Color, colors
from openpyxl.styles import PatternFill
from openpyxl.styles.colors import GREEN

########### BEGIN User Defined Variables 
#
# Make sure to change these variables to match your environment!
# Windows path to racadm.exe.  Please note that you must use forward slashes(/) here!
#CMD = os.path.normpath("C:/Program Files/Dell/SysMgt/idrac/racadm.exe")
# Linux path to racadm binary. 
CMD = "/opt/dell/srvadmin/sbin/racadm"
USERID = "root"
PASSWD = "calvin"
XML_LIST1 = "R730xd_config_V6.xml"
XML_LIST2 = "R630_config_V3.xml"
FNULL = open(os.devnull, 'w')
node_list = subprocess.check_output(["nodels", "imm"]).split()
MODEL1 = "R730xd"
MODEL2 = "R630"
DM = "PowerEdge"
NFS = "192.168.70.254:/dell"
IPLIST = "Header_List.txt"
input_file = open(IPLIST)
# Change to the 3 digit code for the location
LOC = "ash"
# Change to the Rack location
RACK = "r01"
#
########### END User Defined Variables 

FNULL = open(os.devnull, 'w')

wb = Workbook()
ws1 = wb.active
ws1.title = "Ashburn"
dest_filename = 'Gaia_Ashburn_016-020_EDI_Template_Final.xlsx'


i = 2
j = 1
z = 1
for line in input_file:	
	line = line.rstrip('\r\n')
	c = ws1.cell(row = 1, column = j)
	c.value = line
	c.fill = PatternFill(fill_type='solid', start_color='32cd32', end_color='32cd32')
	j = j + 1


def ping(host):
	
	ping_str = "-n 1" if  platform.system().lower()=="windows" else "-c 1"
	args = "ping " + " " + ping_str + " " + host 
	need_sh = False if  platform.system().lower()=="windows" else True

    #Ping
	return subprocess.call(args, shell=need_sh, stdout=FNULL, stderr=subprocess.STDOUT) == 0

for node in node_list:
	node = node.rstrip('\r\n')
	if ping(node):
		#Hardware inventory get
		args = CMD + " -r " + node + " -u " + USERID + " -p " + PASSWD + " get BIOS.SysInformation.systemmodelName | grep 'systemmodelName=' | awk -F= '{print $NF}'"
		argsSN = CMD + " -r " + node + " -u root -p calvin get BIOS.SysInformation.SystemServiceTag | grep 'SystemServiceTag=' | awk -F= '{print $NF}'"
		argsM1 = CMD + " -r " + node + " -u root -p calvin get NIC.VndrConfigPage.1 | grep '#MacAddr=' | awk -F= '{print $NF}'"
		argsM2 = CMD + " -r " + node + " -u root -p calvin get NIC.VndrConfigPage.5 | grep '#MacAddr=' | awk -F= '{print $NF}'"
		argsM3 = CMD + " -r " + node + " -u root -p calvin get NIC.VndrConfigPage.6 | grep '#MacAddr=' | awk -F= '{print $NF}'" 
		argsM4 = CMD + " -r " + node + " -u root -p calvin get IDRAC.NIC | grep 'MACAddress' | awk -F= '{print $NF}'"
		
		need_sh = False if  platform.system().lower()=="windows" else True

		
		try:
			print("Processing " + node + " hardware inventory...")
			rawmodel = subprocess.check_output(args, stderr=subprocess.STDOUT, shell=need_sh).strip('\r\n')
			SerialN = subprocess.check_output(argsSN, stderr=subprocess.STDOUT, shell=need_sh).strip('\r\n')
			MAC1 = subprocess.check_output(argsM1, stderr=subprocess.STDOUT, shell=need_sh).strip('\r\n')
			MAC2 = subprocess.check_output(argsM2, stderr=subprocess.STDOUT, shell=need_sh).strip('\r\n')
			MAC3 = subprocess.check_output(argsM3, stderr=subprocess.STDOUT, shell=need_sh).strip('\r\n')
			MAC4 = subprocess.check_output(argsM4, stderr=subprocess.STDOUT, shell=need_sh).strip('\r\n')
			
			dims = {}
			for row in ws1.rows:
				for cell in row:
					if cell.value:
						dims[cell.column] = max((dims.get(cell.column, 0), len(cell.value)))
			for col, value in dims.items():
				ws1.column_dimensions[col].width = value 
			a = ws1.cell(row = i, column = 1)
			a.value = 'No_Information'
			b = ws1.cell(row = i, column = 2)
			b.value = 'No_Information'
			c = ws1.cell(row = i, column = 3)
			c.value = 'No_Information'
			d = ws1.cell(row = i, column = 4)
			d.value = rawmodel + '  '
			e = ws1.cell(row = i, column = 5)
			e.value = SerialN
			f = ws1.cell(row = i, column = 6)
			f.value = 'No_Information'
			g = ws1.cell(row = i, column = 7)
			g.value = node + '       '
			h = ws1.cell(row = i, column = 8)
			h.value = 'No_Information'
			ii = ws1.cell(row = i, column = 9)
			ii.value = 'No_Information'
			jj = ws1.cell(row = i, column = 10)
			jj.value = 'No_Information'
			k = ws1.cell(row = i, column = 11)
			k.value = 'No_Information'
			l = ws1.cell(row = i, column = 12)
			l.value = 'No_Information'
			m = ws1.cell(row = i, column = 13)
			m.value = 'No_Information'
			n = ws1.cell(row = i, column = 14)
			n.value = MAC1
			o = ws1.cell(row = i, column = 15)
			o.value = MAC2
			p = ws1.cell(row = i, column = 16)
			p.value = MAC3
			q = ws1.cell(row = i, column = 17)
			q.value = MAC4
			r = ws1.cell(row = i, column = 18)
			r.value = 'No_Information'
			s = ws1.cell(row = i, column = 19)
			s.value = 'No_Information'
			t = ws1.cell(row = i, column = 20)
			t.value = 'No_Information'
			u = ws1.cell(row = i, column = 21)
			u.value = 'No_Information'
			v = ws1.cell(row = i, column = 22)
			v.value = 'No_Information'
			i = i + 1
			#z = z + 1
			wb.save(filename = dest_filename)
			print(node + " hardware inventory complete")
			print("Starting automated configuration...")
			rawmodel = subprocess.check_output(args, stderr=subprocess.STDOUT, shell=need_sh)
			if re.findall(MODEL1,rawmodel):
				XML_FILE = XML_LIST1
			elif re.findall(MODEL2,rawmodel):
				XML_FILE = XML_LIST2
			else:
				wmodel = re.match(DM,rawmodel)
				wmodel = rawmodel.split('=')[2]
				print "ERROR: System is not a known model type. Model detected as " + wmodel.rstrip('\r\n') + ".  " + node
				continue
			args = CMD + " -r " + node + " -u " + USERID + " -p " + PASSWD + "jobqueue delete --all" 
			args1 = CMD + " -r " + node + " -u " + USERID + " -p " + PASSWD + " set -t xml -f " + XML_FILE + " -l " + NFS + ' -b "forced"'
			#args2 = CMD + " -r " + line + "-bmc -u " + USERID + " -p " + PASSWD + " set System.Location.DataCenter "  + LOC          
			#args3 = CMD + " -r " + line + "-bmc -u " + USERID + " -p " + PASSWD + " set System.Location.Rack.Name " + RACK 
			#args4 = CMD + " -r " + node + "-bmc -u " + USERID + " -p " + PASSWD + " set System.Location.Rack.Slot " + node
			
			subprocess.call(args, shell=need_sh)	
			#subprocess.call(args2, shell=need_sh)	
			#subprocess.call(args3, shell=need_sh)	
			#subprocess.call(args4, shell=need_sh)	
			time.sleep(15)
			subprocess.call(args1, shell=need_sh)
			print(node + " configuration sent to server")
			print(node + " will be restarting momentarily")
		except subprocess.CalledProcessError as e:
			try:
				print e.output.split("\n")[7] + ": " + node 
			except:
				nond =  e.output.split("\n")
				for error in nond:
					fe = re.search(r'ERROR.*',error)
					if fe:
						print fe.group() + " " + node
	else:
		print "ERROR: Can not ping system " + node + "!!!"
