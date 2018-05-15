from PyQt4 import QtCore, QtGui
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
import subprocess, os, platform, re, time, sys, openpyxl


if not os.path.exists(os.path.expanduser('~\Documents\Inventory')):
    os.makedirs(os.path.expanduser('~\Documents\Inventory'))

save_path = os.path.expanduser('~\Documents\Inventory')
save_filename = 'Inventory and Prices.xlsx'
active_workbook = os.path.join(save_path, save_filename)

if os.path.isfile(save_path + './' + save_filename):
	wb = openpyxl.load_workbook(active_workbook,)
	ws1 = wb.get_sheet_by_name('Log Archive')

else:
	wb = Workbook()
	ws1 = wb.active
	ws1.title = 'Log Archive'
	labels = ['Item Description', 'Item #', 'Quantity', 'P.O Price', 'Sale Price', 'Comments']
	ws1.append(labels)
	ws1.column_dimensions["A"].width = '40'
	ws1.column_dimensions["B"].width = '18'
	ws1.column_dimensions["C"].width = '15'
	ws1.column_dimensions["D"].width = '15'
	ws1.column_dimensions["E"].width = '15'
	ws1.column_dimensions["F"].width = '45'
	wb.save(filename = active_workbook)

try:
	_fromUtf8 = QtCore.QString.fromUtf8
except AttributeError:
	def _fromUtf8(s):
		return s

try:
	_encoding = QtGui.QApplication.UnicodeUTF8
	def _translate(context, text, disambig):
		return QtGui.QApplication.translate(context, text, disambig, _encoding)
except AttributeError:
	def _translate(context, text, disambig):
		return QtGui.QApplication.translate(context, text, disambig)

class Ui_MainWindow(QtGui.QMainWindow):

	def __init__(self):
		QtGui.QMainWindow.__init__(self)
		self.setupUi(self)
		
	def update_table(self):
		filePath = (os.path.expanduser('~\Documents\Inventory./') + 'Inventory and Prices.csv')
		fileHandle = open(filePath, 'r')
		line = fileHandle.read()[:].split(',')
		print line
		for n, val in enumerate(line):
			newitem=QtGui.QTableWidgetItem(val)
			self.cur_inv_tbl.setItem(0, n, newitem)
		self.cur_inv_tbl.resizeColumnsToContents()
		self.cur_inv_tbl.resizeRowsToContents()
		
	def open_directory(self):
		os.startfile(os.path.expanduser('~\Documents\Inventory'))
	
	def submit_data(self):
		item_desc_in_text = str(self.item_desc_in.displayText())
		item_nbr_in_text = str(self.item_nbr_in.displayText())
		qnty_in_text = str(self.qnty_in.displayText())
		po_prc_in_text = str(self.po_prc_in.displayText())
		sale_prc_in_text = str(self.sale_prc_in.displayText())
		comm_in_text = str(self.comm_in.displayText())
		submit_data_entry = (item_desc_in_text, item_nbr_in_text, qnty_in_text, sale_prc_in_text, po_prc_in_text, comm_in_text)
		matching_row_nbr = None
		for rowNum in range(2, ws1.max_row + 1 ):
			log_name = ws1.cell(row=rowNum,column=2).value
			if log_name == item_nbr_in_text:
				matching_row_nbr = rowNum
				break
		if matching_row_nbr is not None:
			if item_desc_in_text is not "":
				ws1.cell(row=matching_row_nbr, column=1).value = item_desc_in_text
				wb.save(filename = active_workbook)
			if qnty_in_text is not "":
				ws1.cell(row=matching_row_nbr, column=3).value = qnty_in_text
				wb.save(filename = active_workbook)
			if po_prc_in_text is not "":
				ws1.cell(row=matching_row_nbr, column=4).value = po_prc_in_text
				wb.save(filename = active_workbook)
			if sale_prc_in_text is not "":
				ws1.cell(row=matching_row_nbr, column=5).value = sale_prc_in_text
				wb.save(filename = active_workbook)
			if comm_in_text is not "":
				ws1.cell(row=matching_row_nbr, column=6).value = comm_in_text
				wb.save(filename = active_workbook)

		else:
			ws1.append(submit_data_entry)
			wb.save(filename = active_workbook)
			self.clear_fields()

	def clear_fields(self):
		self.comm_in.clear()
		self.item_desc_in.clear()
		self.item_nbr_in.clear()
		self.po_prc_in.clear()
		self.qnty_in.clear()
		self.sale_prc_in.clear()
		
	def setupUi(self, MainWindow):
		MainWindow.setObjectName(_fromUtf8("MainWindow"))
		MainWindow.resize(1230, 876)
		font = QtGui.QFont()
		font.setPointSize(10)
		MainWindow.setFont(font)
		self.centralwidget = QtGui.QWidget(MainWindow)
		self.centralwidget.setObjectName(_fromUtf8("centralwidget"))
		self.gridLayout_3 = QtGui.QGridLayout(self.centralwidget)
		self.gridLayout_3.setObjectName(_fromUtf8("gridLayout_3"))
		self.cur_inv_tbl = QtGui.QTableWidget(self.centralwidget)
		self.cur_inv_tbl.setObjectName(_fromUtf8("cur_inv_tbl"))
		self.cur_inv_tbl.setColumnCount(6)
		self.cur_inv_tbl.setRowCount(9)
		item = QtGui.QTableWidgetItem()
		self.cur_inv_tbl.setVerticalHeaderItem(0, item)
		item = QtGui.QTableWidgetItem()
		self.cur_inv_tbl.setVerticalHeaderItem(1, item)
		item = QtGui.QTableWidgetItem()
		self.cur_inv_tbl.setVerticalHeaderItem(2, item)
		item = QtGui.QTableWidgetItem()
		self.cur_inv_tbl.setVerticalHeaderItem(3, item)
		item = QtGui.QTableWidgetItem()
		self.cur_inv_tbl.setVerticalHeaderItem(4, item)
		item = QtGui.QTableWidgetItem()
		self.cur_inv_tbl.setVerticalHeaderItem(5, item)
		item = QtGui.QTableWidgetItem()
		self.cur_inv_tbl.setVerticalHeaderItem(6, item)
		item = QtGui.QTableWidgetItem()
		self.cur_inv_tbl.setVerticalHeaderItem(7, item)
		item = QtGui.QTableWidgetItem()
		self.cur_inv_tbl.setVerticalHeaderItem(8, item)
		item = QtGui.QTableWidgetItem()
		self.cur_inv_tbl.setHorizontalHeaderItem(0, item)
		item = QtGui.QTableWidgetItem()
		self.cur_inv_tbl.setHorizontalHeaderItem(1, item)
		item = QtGui.QTableWidgetItem()
		self.cur_inv_tbl.setHorizontalHeaderItem(2, item)
		item = QtGui.QTableWidgetItem()
		self.cur_inv_tbl.setHorizontalHeaderItem(3, item)
		item = QtGui.QTableWidgetItem()
		self.cur_inv_tbl.setHorizontalHeaderItem(4, item)
		item = QtGui.QTableWidgetItem()
		self.cur_inv_tbl.setHorizontalHeaderItem(5, item)
		self.gridLayout_3.addWidget(self.cur_inv_tbl, 3, 0, 1, 3)
		self.open_btn = QtGui.QPushButton(self.centralwidget)
		self.open_btn.setMinimumSize(QtCore.QSize(115, 32))
		font = QtGui.QFont()
		font.setPointSize(15)
		self.open_btn.setFont(font)
		self.open_btn.setObjectName(_fromUtf8("open_btn"))
		self.open_btn.clicked.connect(self.open_directory)
		self.gridLayout_3.addWidget(self.open_btn, 1, 2, 1, 1)
		self.gridLayout = QtGui.QGridLayout()
		self.gridLayout.setObjectName(_fromUtf8("gridLayout"))
		self.item_desc_lbl = QtGui.QLabel(self.centralwidget)
		self.item_desc_lbl.setMinimumSize(QtCore.QSize(238, 33))
		font = QtGui.QFont()
		font.setPointSize(20)
		font.setBold(True)
		font.setWeight(75)
		self.item_desc_lbl.setFont(font)
		self.item_desc_lbl.setObjectName(_fromUtf8("item_desc_lbl"))
		self.gridLayout.addWidget(self.item_desc_lbl, 0, 0, 1, 1)
		self.item_desc_in = QtGui.QLineEdit(self.centralwidget)
		self.item_desc_in.setMinimumSize(QtCore.QSize(653, 30))
		font = QtGui.QFont()
		font.setPointSize(15)
		self.item_desc_in.setFont(font)
		self.item_desc_in.setObjectName(_fromUtf8("item_desc_in"))
		self.gridLayout.addWidget(self.item_desc_in, 0, 1, 1, 1)
		self.item_nbr_lbl = QtGui.QLabel(self.centralwidget)
		self.item_nbr_lbl.setMinimumSize(QtCore.QSize(238, 33))
		font = QtGui.QFont()
		font.setPointSize(20)
		font.setBold(True)
		font.setWeight(75)
		self.item_nbr_lbl.setFont(font)
		self.item_nbr_lbl.setObjectName(_fromUtf8("item_nbr_lbl"))
		self.gridLayout.addWidget(self.item_nbr_lbl, 1, 0, 1, 1)
		self.item_nbr_in = QtGui.QLineEdit(self.centralwidget)
		self.item_nbr_in.setMinimumSize(QtCore.QSize(653, 30))
		font = QtGui.QFont()
		font.setPointSize(15)
		self.item_nbr_in.setFont(font)
		self.item_nbr_in.setObjectName(_fromUtf8("item_nbr_in"))
		self.gridLayout.addWidget(self.item_nbr_in, 1, 1, 1, 1)
		self.qnty_lbl = QtGui.QLabel(self.centralwidget)
		self.qnty_lbl.setMinimumSize(QtCore.QSize(238, 33))
		font = QtGui.QFont()
		font.setPointSize(20)
		font.setBold(True)
		font.setWeight(75)
		self.qnty_lbl.setFont(font)
		self.qnty_lbl.setObjectName(_fromUtf8("qnty_lbl"))
		self.gridLayout.addWidget(self.qnty_lbl, 2, 0, 1, 1)
		self.qnty_in = QtGui.QLineEdit(self.centralwidget)
		self.qnty_in.setMinimumSize(QtCore.QSize(653, 30))
		font = QtGui.QFont()
		font.setPointSize(15)
		self.qnty_in.setFont(font)
		self.qnty_in.setObjectName(_fromUtf8("qnty_in"))
		self.gridLayout.addWidget(self.qnty_in, 2, 1, 1, 1)
		self.po_prc_lbl = QtGui.QLabel(self.centralwidget)
		self.po_prc_lbl.setMinimumSize(QtCore.QSize(238, 33))
		font = QtGui.QFont()
		font.setPointSize(20)
		font.setBold(True)
		font.setWeight(75)
		self.po_prc_lbl.setFont(font)
		self.po_prc_lbl.setObjectName(_fromUtf8("po_prc_lbl"))
		self.gridLayout.addWidget(self.po_prc_lbl, 3, 0, 1, 1)
		self.po_prc_in = QtGui.QLineEdit(self.centralwidget)
		self.po_prc_in.setMinimumSize(QtCore.QSize(653, 30))
		font = QtGui.QFont()
		font.setPointSize(15)
		self.po_prc_in.setFont(font)
		self.po_prc_in.setObjectName(_fromUtf8("po_prc_in"))
		self.gridLayout.addWidget(self.po_prc_in, 3, 1, 1, 1)
		self.sale_prc_lbl = QtGui.QLabel(self.centralwidget)
		self.sale_prc_lbl.setMinimumSize(QtCore.QSize(238, 33))
		font = QtGui.QFont()
		font.setPointSize(20)
		font.setBold(True)
		font.setWeight(75)
		self.sale_prc_lbl.setFont(font)
		self.sale_prc_lbl.setCursor(QtGui.QCursor(QtCore.Qt.ArrowCursor))
		self.sale_prc_lbl.setObjectName(_fromUtf8("sale_prc_lbl"))
		self.gridLayout.addWidget(self.sale_prc_lbl, 4, 0, 1, 1)
		self.sale_prc_in = QtGui.QLineEdit(self.centralwidget)
		self.sale_prc_in.setMinimumSize(QtCore.QSize(653, 30))
		font = QtGui.QFont()
		font.setPointSize(15)
		self.sale_prc_in.setFont(font)
		self.sale_prc_in.setObjectName(_fromUtf8("sale_prc_in"))
		self.gridLayout.addWidget(self.sale_prc_in, 4, 1, 1, 1)
		self.comm_lbl = QtGui.QLabel(self.centralwidget)
		self.comm_lbl.setMinimumSize(QtCore.QSize(238, 33))
		font = QtGui.QFont()
		font.setPointSize(20)
		font.setBold(True)
		font.setWeight(75)
		self.comm_lbl.setFont(font)
		self.comm_lbl.setObjectName(_fromUtf8("comm_lbl"))
		self.gridLayout.addWidget(self.comm_lbl, 5, 0, 1, 1)
		self.comm_in = QtGui.QLineEdit(self.centralwidget)
		self.comm_in.setMinimumSize(QtCore.QSize(653, 30))
		font = QtGui.QFont()
		font.setPointSize(15)
		self.comm_in.setFont(font)
		self.comm_in.setObjectName(_fromUtf8("comm_in"))
		self.gridLayout.addWidget(self.comm_in, 5, 1, 1, 1)
		self.gridLayout_3.addLayout(self.gridLayout, 0, 0, 1, 3)
		spacerItem = QtGui.QSpacerItem(533, 34, QtGui.QSizePolicy.Expanding, QtGui.QSizePolicy.Minimum)
		self.gridLayout_3.addItem(spacerItem, 1, 1, 1, 1)
		self.gridLayout_2 = QtGui.QGridLayout()
		self.gridLayout_2.setObjectName(_fromUtf8("gridLayout_2"))
		self.clear_btn = QtGui.QPushButton(self.centralwidget)
		self.clear_btn.setMinimumSize(QtCore.QSize(75, 32))
		font = QtGui.QFont()
		font.setPointSize(15)
		self.clear_btn.setFont(font)
		self.clear_btn.setObjectName(_fromUtf8("clear_btn"))
		self.clear_btn.clicked.connect(self.update_table)
		self.gridLayout_2.addWidget(self.clear_btn, 0, 1, 1, 1)
		self.submit_btn = QtGui.QPushButton(self.centralwidget)
		self.submit_btn.setMinimumSize(QtCore.QSize(75, 32))
		font = QtGui.QFont()
		font.setPointSize(15)
		self.submit_btn.setFont(font)
		self.submit_btn.setObjectName(_fromUtf8("submit_btn"))
		self.submit_btn.clicked.connect(self.submit_data)
		self.gridLayout_2.addWidget(self.submit_btn, 0, 0, 1, 1)
		self.close_btn = QtGui.QPushButton(self.centralwidget)
		self.close_btn.setMinimumSize(QtCore.QSize(75, 32))
		font = QtGui.QFont()
		font.setPointSize(15)
		self.close_btn.setFont(font)
		self.close_btn.setObjectName(_fromUtf8("close_btn"))
		self.close_btn.clicked.connect(QtCore.QCoreApplication.instance().quit)
		self.gridLayout_2.addWidget(self.close_btn, 0, 2, 1, 2)
		self.gridLayout_3.addLayout(self.gridLayout_2, 1, 0, 1, 1)
		self.cur_inv_lbl = QtGui.QLabel(self.centralwidget)
		font = QtGui.QFont()
		font.setPointSize(20)
		font.setBold(True)
		font.setWeight(75)
		self.cur_inv_lbl.setFont(font)
		self.cur_inv_lbl.setObjectName(_fromUtf8("cur_inv_lbl"))
		self.gridLayout_3.addWidget(self.cur_inv_lbl, 2, 0, 1, 1)
		MainWindow.setCentralWidget(self.centralwidget)
		self.menubar = QtGui.QMenuBar(MainWindow)
		self.menubar.setGeometry(QtCore.QRect(0, 0, 1230, 23))
		self.menubar.setObjectName(_fromUtf8("menubar"))
		MainWindow.setMenuBar(self.menubar)
		self.statusbar = QtGui.QStatusBar(MainWindow)
		self.statusbar.setObjectName(_fromUtf8("statusbar"))
		MainWindow.setStatusBar(self.statusbar)

		self.retranslateUi(MainWindow)
		QtCore.QMetaObject.connectSlotsByName(MainWindow)

	def retranslateUi(self, MainWindow):
		MainWindow.setWindowTitle(_translate("MainWindow", "MainWindow", None))
		item = self.cur_inv_tbl.verticalHeaderItem(0)
		item.setText(_translate("MainWindow", "1", None))
		item = self.cur_inv_tbl.verticalHeaderItem(1)
		item.setText(_translate("MainWindow", "2", None))
		item = self.cur_inv_tbl.verticalHeaderItem(2)
		item.setText(_translate("MainWindow", "3", None))
		item = self.cur_inv_tbl.verticalHeaderItem(3)
		item.setText(_translate("MainWindow", "4", None))
		item = self.cur_inv_tbl.verticalHeaderItem(4)
		item.setText(_translate("MainWindow", "5", None))
		item = self.cur_inv_tbl.verticalHeaderItem(5)
		item.setText(_translate("MainWindow", "6", None))
		item = self.cur_inv_tbl.verticalHeaderItem(6)
		item.setText(_translate("MainWindow", "7", None))
		item = self.cur_inv_tbl.verticalHeaderItem(7)
		item.setText(_translate("MainWindow", "8", None))
		item = self.cur_inv_tbl.verticalHeaderItem(8)
		item.setText(_translate("MainWindow", "10", None))
		item = self.cur_inv_tbl.horizontalHeaderItem(0)
		item.setText(_translate("MainWindow", "A", None))
		item = self.cur_inv_tbl.horizontalHeaderItem(1)
		item.setText(_translate("MainWindow", "B", None))
		item = self.cur_inv_tbl.horizontalHeaderItem(2)
		item.setText(_translate("MainWindow", "C", None))
		item = self.cur_inv_tbl.horizontalHeaderItem(3)
		item.setText(_translate("MainWindow", "D", None))
		item = self.cur_inv_tbl.horizontalHeaderItem(4)
		item.setText(_translate("MainWindow", "E", None))
		item = self.cur_inv_tbl.horizontalHeaderItem(5)
		item.setText(_translate("MainWindow", "F", None))
		self.open_btn.setText(_translate("MainWindow", "Open Folder", None))
		self.item_desc_lbl.setText(_translate("MainWindow", "Item Description:", None))
		self.item_nbr_lbl.setText(_translate("MainWindow", "Item #:", None))
		self.qnty_lbl.setText(_translate("MainWindow", "Quantity:", None))
		self.po_prc_lbl.setText(_translate("MainWindow", "P.O Price:", None))
		self.sale_prc_lbl.setText(_translate("MainWindow", "Sale Price:", None))
		self.comm_lbl.setText(_translate("MainWindow", "Comments:", None))
		self.clear_btn.setText(_translate("MainWindow", "Clear", None))
		self.submit_btn.setText(_translate("MainWindow", "Submit", None))
		self.close_btn.setText(_translate("MainWindow", "Close", None))
		self.cur_inv_lbl.setText(_translate("MainWindow", "Current Inventory", None))

if __name__ == '__main__':
	app = QtGui.QApplication(sys.argv)
	window = Ui_MainWindow()
	window.show()
	sys.exit(app.exec_())