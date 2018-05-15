import os
import subprocess
import re
import sys
import argparse
from openpyxl import load_workbook, Workbook
import warnings

import KitsuUtils.core as Core

FNULL = open(os.devnull, 'w')
warnings.filterwarnings("ignore")

def OpenFile(path):

    """ 
    Opens Project BOM file.
    
    Args:
        path: Path to the file to be opened.
    
    Returns:
        Workbook object from file.
    """

    try:
        return load_workbook(path, data_only=True)
    except:
        print "File does not exist!"
        sys.exit(1)

def ConvertName(orig_name):

    """
    Returns node name as used in Kitsu from name in BOM

    Args:
        orig_name: Name from BOM file

    Returns:
        String: converted name
    """

    name = orig_name.replace('cld', '').lstrip('0')
    return 'cld{}'.format(name)

class Node(object):

    """ 
    Represents a generic node 
    
    """

    ROLE_TO_MODEL = {
        'mgmt': 'R630',
        'compute': 'R630',
        'storage': 'R730xd',
        'gpu': 'C4130'
    }

    connections = []
        
    def __init__(self, sitecode, deviceName, deviceRole, deviceConnection = None):

        self.name = deviceName
        self.name = self.name.replace('-', '').replace('cld', 'gaia-r')
        self.bmc = self.name + "-bmc"
        self.rack = re.search('(?<=gaia-r)([0-9]+)', self.name).group(0)
        self.u = self.name[-2:].lstrip('0')
        self.connections.append(deviceConnection)
        self.gigIP = "10.2.%s.%s" % (self.rack, self.u)
        self.bmcIP = "10.2.%s.%s" % (self.rack, int(self.u) + 100)
        self.role = deviceRole.rstrip('server').strip()
        self.model = Node.ROLE_TO_MODEL[self.role]
        self.pool = sitecode.lower()
        self.groups = "dell"
        self.groups += ",gaia-node"
        self.groups += ",gaia-{}".format(deviceRole.rstrip('server').strip())
        self.groups += ",gaia-{}-node".format(self.pool)
        self.groups += ',gaia-{}-{}'.format(self.pool, deviceRole.rstrip('server').strip())
        self.groups += ",gaia-r{}".format(self.rack)
        self.groups += ',gaia-r{}-{}'.format(self.rack, deviceRole.rstrip('server').strip())
        self.definition = ''
        self.bmcdefinition = ''


class Switch(object):

    """ Represents a network switch """

    connections = []
        
    def __init__(self, sitecode, deviceName, deviceRole, deviceModel, deviceConnection = None):
        self.name = deviceName.replace('-', '').replace('cld', 'gaia-r')
        
        self.model = deviceModel
        
        self.role = deviceRole
        if not 'spare' in self.role:
            self.isSpare = False
        else:
            self.isSpare = True

        self.rack = re.search('(?<=gaia-r)([0-9]+)', self.name).group(0)
        self.u = self.name[-2:]
        self.connections.append(deviceConnection)
        self.definition = ''
        self.pool = sitecode.lower()

        if self.role == 'console-srv':
            self.groups = 'switch,serial'
        else:
            self.groups = 'switch,cumulus,gaia-sw'
            self.groups += ',gaia-r{}-sw'.format(self.rack)
            self.groups += ',gaia-{}-sw'.format(self.pool)
            self.groups += ',gaia-{}-{}'.format(self.pool, self.role)
            self.groups += ',gaia-r{}-{}'.format(self.rack, self.role)           

    def GetDescription(self):
        return '[%s] Rack: %s U: %s' % (self.name, self.rack, self.u)

    def __str__(self):
        return '[%s] Rack: %s U: %s' % (self.name, self.rack, self.u)


# Represents a device to device connection
class Connection(object):
    def __init__(self, aDeviceName, aDeviceRole, aDeviceModel, aDevicePort, aDeviceXcvr, bDeviceName, bDeviceRole, bDeviceModel, bDevicePort, bDeviceXcvr, connectionRole, cableComments):
        self.aDeviceName = aDeviceName
        
        if not 'spare' in aDeviceRole:
            self.aDeviceRole = aDeviceRole[0:-3]
        else:
            self.aDeviceRole = aDeviceRole

        self.aDeviceModel = aDeviceModel
        self.aDevicePort = aDevicePort
        self.aDeviceXcvr = aDeviceXcvr
        self.bDeviceName = bDeviceName
        
        if not 'spare' in bDeviceRole:
            self.bDeviceRole = bDeviceRole[0:-3]
        else:
            self.bDeviceRole = bDeviceRole

        self.bDeviceModel = bDeviceModel
        self.bDevicePort = bDevicePort
        self.bDeviceXcvr = bDeviceXcvr
        self.connectionRole = connectionRole
        self.cableComments = cableComments

        self.aDeviceName = ConvertName(self.aDeviceName)
        self.bDeviceName = ConvertName(self.bDeviceName)


def GetSiteCode(bom_filename):
    return bom_filename[0:5]

# Gets the cell in which connection data starts from the Project BOM
def GetDataStart(ws):
    for row in ws.iter_rows():
        for cell in row:
            if cell.value == 'DEVICE_A_NAME':
                return cell.row + 1


def GetConnectionData(ws):

    """
    Parses connection data from Project BOM.

    Args:
        wb - The workbook to be parsed.
    """
    connections = []

    row = GetDataStart(ws)
    while row <= ws.max_row:
        if ws['A%i' % row].value != None and 'cld' in ws['A%i' % row].value:
            if not ws.row_dimensions[row].hidden:
                connections.append(Connection(ws['A%i' % row].value, ws['C%i' % row].value, ws['D%i' % row].value, ws['E%i' % row].value, ws['F%i' % row].value, ws['G%i' % row].value,
                                            ws['I%i' % row].value, ws['J%i' % row].value, ws['K%i' % row].value, ws['L%i' % row].value, ws['M%i' % row].value, ws['N%i' % row].value))
        row += 1

    return connections

def GenerateSwitchDefinitions(sitecode, connections, agg_port):

    """
    Generates switch definitions for xCAT import into a .txt file in the current directory.
    
    """

    switches = {}

    for connection in connections:        
        if not connection.aDeviceName in switches and (connection.aDeviceRole == 'oobagg' or connection.bDeviceRole == 'oobagg'):
            switches[connection.aDeviceName] = Switch(sitecode, connection.aDeviceName, connection.aDeviceRole, connection.aDeviceModel, connection)

            if switches[connection.aDeviceName].isSpare or connection.aDeviceRole == 'console-srv':
                continue

            definition = "mkdef %s --template onieswitch arch=armv71 " % switches[connection.aDeviceName].name
            definition += "groups=%s " % switches[connection.aDeviceName].groups
            
            if connection.aDeviceRole == 'oobagg':                
                definition += "ip=10.2.%s.0 " % switches[connection.aDeviceName].rack
                definition += "switch=br1coresw1 "
                definition += "switchport=199/1/%s " % agg_port

            elif connection.connectionRole == 'mgmt_conn':
                if connection.aDeviceRole == 'bmcmgmt':
                    definition += "ip=10.2.%s.255 " % switches[connection.aDeviceName].rack                    
                elif connection.aDeviceRole == 'leaf':
                    definition += "ip=10.2.%s.254 " % switches[connection.aDeviceName].rack
                elif connection.aDeviceRole == 'spine':
                    definition += "ip=10.2.%s.253 " % switches[connection.aDeviceName].rack
                elif connection.aDeviceRole == 'border':
                    definition += "ip=10.2.%s.252 " % switches[connection.aDeviceName].rack
                elif connection.aDeviceRole == 'span-agg':
                    definition += "ip=10.2.%s.251 " % switches[connection.aDeviceName].rack
                elif connection.aDeviceRole == 'console-srv':
                    definition += "ip=10.2.%s.250 " % switches[connection.aDeviceName].rack
                
                definition += "switch=%s " % switches[connection.bDeviceName].name
                definition += "switchport=swp%s " % connection.bDevicePort
           
            definition += "provmethod=/install/custom/sw_os/cumulus/CumulusLinux-2.5.12-amd64.bin "

            switches[connection.aDeviceName].definition = definition

        elif connection.aDeviceName in switches:
            switches[connection.aDeviceName].connections.append(connection)     

    return switches       


def GenerateNodeDefinitions(sitecode, connections, switches):

    """
    Generates node definitions for xCAT import into a .txt file in the current directory.

    """

    nodes = {}
    
    for connection in connections:

        definition = 'chdef '
        bmcdefinition = 'chdef '

        if not connection.aDeviceName in nodes and not connection.aDeviceName in switches:
            nodes[connection.aDeviceName] = Node(sitecode, connection.aDeviceName, connection.aDeviceRole, connection)    

        if connection.connectionRole == 'bmc_conn':
            definition += "%s " % nodes[connection.aDeviceName].name
            definition += "ip=%s " % nodes[connection.aDeviceName].gigIP
            definition += "groups=%s " % nodes[connection.aDeviceName].groups
            definition += "mgt=ipmi "
            definition += "bmc=%s " % nodes[connection.aDeviceName].bmc
            definition += "bmcport=1 "
            definition += "switch=%s " % switches[connection.bDeviceName].name
            definition += "switchport=swp%s " % connection.bDevicePort
            definition += "usercomment=%s " % nodes[connection.aDeviceName].role

            nodes[connection.aDeviceName].definition = definition

            bmcdefinition += "%s-bmc " % nodes[connection.aDeviceName].name
            bmcdefinition += "ip=%s " % nodes[connection.aDeviceName].bmcIP

            bmcgroups = "gaia-r{}-bmc".format(nodes[connection.aDeviceName].rack)
            bmcgroups += ',gaia-{}-bmc'.format(nodes[connection.aDeviceName].pool)
            bmcgroups += ',gaia-bmc'
            bmcdefinition += "groups=%s " % bmcgroups                         

            nodes[connection.aDeviceName].bmcdefinition = bmcdefinition

        if connection.bDeviceName in switches:
            if not connection in switches[connection.bDeviceName].connections:
                switches[connection.bDeviceName].connections.append(connection)

    return nodes, switches


def GeneratePoolDict(switches, nodes):
    
    pool_dict = {}

    for k, switch in switches.items():
        rack_name = 'gaia-r' + switch.rack
        if not rack_name in pool_dict:            
            pool_dict[rack_name] = { 'switches': [], 'nodes': [] } 
        pool_dict[rack_name]['switches'].append(switch)

    for k, node in nodes.items():
        rack_name = 'gaia-r' + node.rack
        pool_dict[rack_name]['nodes'].append(node)

    return pool_dict


def WriteDefinitions(sitecode, switches, nodes):
    
    """
    Writes switch and node definitions to .txt files in the current directory.

    """

    definitions = []

    for k, switch in switches.items():
        if switch.definition:
            definitions.append(switch.definition)
  
    for k, node in nodes.items():
        definitions.append(node.definition)
        definitions.append(node.bmcdefinition)

    return definitions

def generate_pool(filename, agg_port, inject=False):
    
    wb = OpenFile(filename)
    sitecode = GetSiteCode(os.path.basename(filename))
    connections = GetConnectionData(wb.get_sheet_by_name('Network Fabric'))
    switches = GenerateSwitchDefinitions(sitecode, connections, agg_port)

    cabinetSheets = ('cabinet1', 'cabinet2', 'cabinet3', 'cabinet4', 'cabinet5')
    for sheet in cabinetSheets:
        ws = wb.get_sheet_by_name(sheet)
        connections += GetConnectionData(ws)
    nodes, switches = GenerateNodeDefinitions(sitecode, connections, switches)

    pool_dict = GeneratePoolDict(switches, nodes)
    definition_list = WriteDefinitions(sitecode, switches, nodes)
    
    if inject:
        create_definitions(definition_list)

    return pool_dict

def inject_definition(node_definition):
    try:
        return subprocess.check_output(node_definition, shell=True, stderr=subprocess.STDOUT)
    except Exception as e:
        print e

def create_definitions(definition_list):
    
    work_list = []
    for definition in definition_list:
        func = inject_definition
        fargs = ('/opt/xcat/bin/' + definition, )
        work_list.append((func, fargs))

    output = []
    for result in Core.p_execute(work_list):
        if result:
            output.append(result)
            
    return output    
    

if __name__ == '__main__':    
    parser = argparse.ArgumentParser(description='GAIA - Pool Definition Generator')
    parser.add_argument('filename', help='Path to the project BOM to parse (required).')
    parser.add_argument("-i", "--inject", action='store_true', help='Inject definitions directly into Kitsu')
    
    args = parser.parse_args()
    agg_port = raw_input('Nexus switchport for OOB-AGG (ex. 199/1/X): ')    
    definition_list = generate_pool(args.filename, agg_port, args.inject)
        
    